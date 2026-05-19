#!/usr/bin/env python3
"""
Script diario: descarga el Excel de Google Drive y sube los datos a Supabase.
Uso: python3 upload_to_supabase.py ruta/al/archivo.xlsx
"""

import sys
import json
import argparse
from datetime import datetime
import openpyxl
import requests

# ── CONFIG ────────────────────────────────────────────────────────────────────
SUPABASE_URL = "https://ctbkipgraypohrdeoggv.supabase.co"
SUPABASE_KEY = "sb_publishable_xcIKLzG17KzKOUHA1QlpoQ_WHWnz95d"
PERIODO      = "2026-04-01_2026-05-07"   # actualizar cada corrida

HEADERS = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "resolution=merge-duplicates",
}

# ── PARSE ─────────────────────────────────────────────────────────────────────
def parse_date(val):
    if not val:
        return None
    s = str(val)[:10]
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None

def parse_hora_llegada(val):
    """Extract YYYY-MM-DD from 'DD/MM/YYYY, HH:MM a.m./p.m.' (always Mexico time)."""
    if not val:
        return None
    date_part = str(val).split(',')[0].strip()
    return parse_date(date_part)

def find_header_row(ws):
    """Find the row index that contains 'ID candidato'."""
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if any(str(v) == "ID candidato" for v in row if v):
            return i
    raise ValueError("No se encontró la fila de encabezados con 'ID candidato'")

def load_excel(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Funnel"]
    rows = list(ws.iter_rows(values_only=True))
    hdr_idx = find_header_row(ws)
    headers = [str(h) if h else f"col_{i}" for i, h in enumerate(rows[hdr_idx])]
    records = []
    for row in rows[hdr_idx + 1:]:
        if all(v is None for v in row):
            continue
        r = dict(zip(headers, row))
        records.append({
            "id_candidato":        str(r.get("ID candidato") or ""),
            "candidato":           r.get("Candidato"),
            "canal":               r.get("Canal"),
            "inicio":              parse_date(r.get("Inicio")),
            "hora_llegada":        parse_hora_llegada(r.get("Hora de llegada")),
            "asistencia":          r.get("Asistencia"),
            "declina_oferta":      r.get("Declina oferta"),
            "razon_declina":       r.get("Razón declina oferta"),
            "entrevistador":       r.get("Entrevistador/a"),
            "estado_entrevista":   r.get("Estado entrevista"),
            "razon_rechazo":       r.get("Razón de rechazo"),
            "background_check":    r.get("Background check"),
            "suma_bg":             r.get("La suma de entrevista y background"),
            "prueba_manejo":       r.get("Prueba de manejo"),
            "razon_no_aprobado":   r.get("Razón prueba de manejo") or r.get("Razón de no aprobado"),
            "onboarding_dia":      r.get("Día asignado onboarding"),
            "asistencia_onboarding": r.get("Asistencia de onboarding"),
            "docs_y_didi":         r.get("Docs y DiDi"),
            "status_driver":       r.get("Status Driver"),
            "firma_contrato":      parse_date(r.get("Firma de contrato")),
            "ultima_modificacion": str(r.get("Última modificación") or ""),
            "periodo":             PERIODO,
        })
    # Deduplicate by id_candidato, keeping last occurrence
    seen = {}
    for r in records:
        seen[r["id_candidato"]] = r
    deduped = list(seen.values())
    if len(deduped) < len(records):
        print(f"  (deduplicados: {len(records) - len(deduped)} filas eliminadas)")
    return deduped

# ── UPLOAD ────────────────────────────────────────────────────────────────────
def upload(records, batch=200):
    total = len(records)
    uploaded = 0
    skip_fields = set()

    for i in range(0, total, batch):
        chunk = records[i:i+batch]
        if skip_fields:
            chunk = [{k: v for k, v in r.items() if k not in skip_fields} for r in chunk]

        resp = requests.post(
            f"{SUPABASE_URL}/rest/v1/funnel",
            headers=HEADERS,
            json=chunk,
        )

        # If a column doesn't exist yet, strip it and retry this batch
        if resp.status_code == 400:
            try:
                err = resp.json()
            except Exception:
                err = {}
            if err.get('code') == 'PGRST204':
                import re
                col = re.search(r"'(\w+)' column", err.get('message', ''))
                if col:
                    field = col.group(1)
                    skip_fields.add(field)
                    print(f"  ⚠ Columna '{field}' no existe en Supabase, se omitirá hasta que se agregue")
                    chunk = [{k: v for k, v in r.items() if k not in skip_fields} for r in chunk]
                    resp = requests.post(
                        f"{SUPABASE_URL}/rest/v1/funnel",
                        headers=HEADERS,
                        json=chunk,
                    )

        if resp.status_code not in (200, 201):
            print(f"  ERROR en batch {i}: {resp.status_code} {resp.text[:200]}")
        else:
            uploaded += len(chunk)
            print(f"  ✓ {uploaded}/{total} registros subidos")

    if skip_fields:
        print(f"  ℹ Campos omitidos (columna faltante en Supabase): {', '.join(skip_fields)}")
    return uploaded

# ── MAIN ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx", help="Ruta al archivo Excel (.xlsx)")
    parser.add_argument("--periodo", help="Periodo (ej: 2026-05-01_2026-05-08)", default=PERIODO)
    args = parser.parse_args()

    PERIODO = args.periodo

    print(f"Leyendo {args.xlsx}...")
    records = load_excel(args.xlsx)
    print(f"  {len(records)} candidatos encontrados")

    print("Subiendo a Supabase...")
    n = upload(records)
    print(f"\nListo. {n} registros en Supabase.")
